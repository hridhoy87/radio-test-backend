from contextlib import asynccontextmanager
from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import desc
import logging
from typing import List, Optional
import time
from sqlalchemy import func
from datetime import datetime, date
from typing import Optional
from . import crud, models, schemas, database
from .database import engine, get_db, Base
from .station_report_routes import router as station_report_router
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from fastapi.responses import JSONResponse
from fastapi.requests import Request
from contextlib import asynccontextmanager
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import logging
import time

from .database import engine, Base
from .station_report_routes import router as station_report_router

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Lifespan manager
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup - just log connection status
    try:
        # Test connection instead of creating tables
        with engine.connect() as conn:
            logger.info("✅ Database connection successful!")
    except Exception as e:
        logger.error(f"❌ Database connection failed: {e}")
    
    yield  # App runs here
    
    # Shutdown
    logger.info("🛑 Application shutting down...")

app = FastAPI(
    title="Radio Test Backend API",
    description="Backend service for collecting radio test location data",
    version="1.0.0",
    lifespan=lifespan
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include routers
app.include_router(station_report_router)

# ==================== BASIC ENDPOINTS ====================

@app.get("/")
async def root():
    return {"message": "Radio Test Backend API", "status": "healthy"}

@app.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": time.time()}

# [Keep your existing endpoints for bulk upload, samples, debug, etc.]
# ... your existing code for other endpoints ...

# ==================== ADVANCED REPORT ENDPOINTS ====================

def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two points using Haversine formula"""
    R = 6371000  # Earth radius in meters
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = (math.sin(delta_lat/2) * math.sin(delta_lat/2) + 
         math.cos(lat1_rad) * math.cos(lat2_rad) * 
         math.sin(delta_lon/2) * math.sin(delta_lon/2))
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c

def time_difference(time1, time2):
    """Calculate time difference in seconds between two HH:MM:SS strings"""
    h1, m1, s1 = map(int, time1.split(':'))
    h2, m2, s2 = map(int, time2.split(':'))
    
    total_seconds1 = h1 * 3600 + m1 * 60 + s1
    total_seconds2 = h2 * 3600 + m2 * 60 + s2
    
    return abs(total_seconds1 - total_seconds2)

def get_comm_state_value(comm_state):
    """Convert comm_state string to numeric value"""
    comm_state = str(comm_state).strip().lower()
    if 'loud and clear' in comm_state:
        return 3
    elif 'readable noisy' in comm_state:
        return 2
    elif 'noisy' in comm_state:
        return 1
    else:
        return 0

def get_comm_state_style(comm_state_value):
    """Get Excel style based on comm_state value"""
    if comm_state_value == 3:
        return PatternFill(start_color="006400", end_color="006400", fill_type="solid"), Font(color="FFFFFF")  # Deep green, white text
    elif comm_state_value == 2:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"), Font(color="000000")  # Light green, black text
    elif comm_state_value == 1:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"), Font(color="000000")  # Orange, black text
    else:
        return PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"), Font(color="FFFFFF")  # Blood red, white text

# ===============Generate station report================

app.include_router(station_report_router)

# ==================== BULK UPLOAD ENDPOINT ====================
@app.post("/locations/bulk", status_code=status.HTTP_201_CREATED)
async def bulk_upload_locations(
    request: schemas.BulkUploadRequest,
    db: Session = Depends(get_db)
):
    """
    Bulk upload location samples from mobile devices.
    """
    try:
        logger.info(f"Received bulk upload from device {request.deviceId} with {len(request.samples)} samples")
        
        # Validate sample count
        if len(request.samples) > 1000:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Too many samples. Maximum 1000 per request, got {len(request.samples)}"
            )
        
        # Process the bulk insert
        successful_ids, failed_samples = crud.create_location_samples_bulk(
            db, request.samples, request.deviceId
        )
        
        # Prepare response
        response = {
            "status": "success",
            "message": f"Processed {len(successful_ids)} samples successfully, {len(failed_samples)} failed",
            "synced_ids": successful_ids,
            "timestamp": time.time()
        }
        
        # Log results
        if failed_samples:
            logger.warning(f"Some samples failed: {failed_samples}")
        
        logger.info(f"Bulk upload completed for device {request.deviceId}: {len(successful_ids)} successful")
        
        return response
        
    except Exception as e:
        logger.error(f"Unexpected error in bulk upload: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Internal server error occurred"
        )

# ==================== GET ENDPOINTS (VIEW DATA) ====================

@app.get("/samples")
async def get_all_samples(
    skip: int = 0, 
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """
    Get all location samples with pagination
    """
    try:
        samples = db.query(models.LocationSample)\
                   .order_by(desc(models.LocationSample.received_at))\
                   .offset(skip)\
                   .limit(limit)\
                   .all()
        
        # Convert to dict to avoid Pydantic issues
        result = []
        for sample in samples:
            result.append({
                "server_id": str(sample.server_id),
                "client_id": sample.client_id,
                "device_id": sample.device_id,
                "lat": sample.lat,
                "lon": sample.lon,
                "acc": sample.acc,
                "sample_date": sample.sample_date,
                "sample_time": sample.sample_time,
                "provider": sample.provider,
                "freq": sample.freq,
                "rf_pwr": sample.rf_pwr,
                "comm_state": sample.comm_state,
                "user": sample.user,
                "station": sample.station,
                "captured_at_utc": sample.captured_at_utc,
                "received_at": sample.received_at.isoformat() if sample.received_at else None,
                "processed": sample.processed,
                "synced_at_utc": sample.synced_at_utc,
            })
        
        return {
            "total": len(result),
            "skip": skip,
            "limit": limit,
            "samples": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/samples/count")
async def get_samples_count(db: Session = Depends(get_db)):
    """
    Get total count of samples in database
    """
    try:
        count = db.query(models.LocationSample).count()
        return {"total_samples": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/samples/device/{device_id}")
async def get_samples_by_device(
    device_id: str,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """
    Get samples by device ID
    """
    try:
        samples = db.query(models.LocationSample)\
                   .filter(models.LocationSample.device_id == device_id)\
                   .order_by(desc(models.LocationSample.received_at))\
                   .offset(skip)\
                   .limit(limit)\
                   .all()
        
        # Convert to dict
        result = []
        for sample in samples:
            result.append({
                "server_id": str(sample.server_id),
                "client_id": sample.client_id,
                "device_id": sample.device_id,
                "lat": sample.lat,
                "lon": sample.lon,
                "acc": sample.acc,
                "sample_date": sample.sample_date,
                "sample_time": sample.sample_time,
                "provider": sample.provider,
                "freq": sample.freq,
                "rf_pwr": sample.rf_pwr,
                "comm_state": sample.comm_state,
                "user": sample.user,
                "station": sample.station,
                "captured_at_utc": sample.captured_at_utc,
                "received_at": sample.received_at.isoformat() if sample.received_at else None,
                "processed": sample.processed
            })
        
        return {
            "device_id": device_id,
            "total_found": len(result),
            "skip": skip,
            "limit": limit,
            "samples": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/samples/{client_id}")
async def get_sample_by_id(client_id: str, db: Session = Depends(get_db)):
    """
    Get a specific sample by client_id
    """
    try:
        sample = db.query(models.LocationSample)\
                  .filter(models.LocationSample.client_id == client_id)\
                  .first()
        
        if not sample:
            raise HTTPException(status_code=404, detail="Sample not found")
        
        return {
            "server_id": str(sample.server_id),
            "client_id": sample.client_id,
            "device_id": sample.device_id,
            "lat": sample.lat,
            "lon": sample.lon,
            "acc": sample.acc,
            "sample_date": sample.sample_date,
            "sample_time": sample.sample_time,
            "provider": sample.provider,
            "freq": sample.freq,
            "rf_pwr": sample.rf_pwr,
            "comm_state": sample.comm_state,
            "user": sample.user,
            "station": sample.station,
            "captured_at_utc": sample.captured_at_utc,
            "received_at": sample.received_at.isoformat() if sample.received_at else None,
            "processed": sample.processed
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

# ==================== DEBUG ENDPOINTS ====================

@app.get("/debug/tables")
async def debug_tables(db: Session = Depends(get_db)):
    """
    Debug endpoint to check if tables exist and their structure
    """
    try:
        # Check if table exists and get sample data
        sample = db.query(models.LocationSample).first()
        
        if sample:
            return {
                "status": "table_exists",
                "table_name": "location_samples",
                "sample_record": {
                    "client_id": sample.client_id,
                    "device_id": sample.device_id,
                    "lat": sample.lat,
                    "lon": sample.lon,
                    "received_at": sample.received_at.isoformat() if sample.received_at else None
                },
                "total_records": db.query(models.LocationSample).count()
            }
        else:
            return {
                "status": "table_exists_but_empty",
                "table_name": "location_samples", 
                "total_records": 0
            }
            
    except Exception as e:
        return {
            "status": "error",
            "message": f"Table may not exist: {str(e)}"
        }

@app.get("/debug/full-sample/{client_id}")
async def debug_full_sample(client_id: str, db: Session = Depends(get_db)):
    """
    Get complete sample data including all fields
    """
    try:
        sample = db.query(models.LocationSample)\
                  .filter(models.LocationSample.client_id == client_id)\
                  .first()
        
        if not sample:
            raise HTTPException(status_code=404, detail="Sample not found")
        
        # Return as dict to avoid Pydantic issues
        return {
            "server_id": str(sample.server_id),
            "client_id": sample.client_id,
            "lat": sample.lat,
            "lon": sample.lon,
            "acc": sample.acc,
            "sample_date": sample.sample_date,
            "sample_time": sample.sample_time,
            "provider": sample.provider,
            "freq": sample.freq,
            "rf_pwr": sample.rf_pwr,
            "comm_state": sample.comm_state,
            "user": sample.user,
            "station": sample.station,
            "captured_at_utc": sample.captured_at_utc,
            "device_id": sample.device_id,
            "received_at": sample.received_at.isoformat() if sample.received_at else None,
            "processed": sample.processed
        }
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/all-fields")
async def debug_all_fields(db: Session = Depends(get_db)):
    """
    Get first sample with all fields for debugging
    """
    try:
        sample = db.query(models.LocationSample).first()
        if sample:
            return {
                "server_id": str(sample.server_id),
                "client_id": sample.client_id,
                "lat": sample.lat,
                "lon": sample.lon,
                "acc": sample.acc,
                "sample_date": sample.sample_date,
                "sample_time": sample.sample_time,
                "provider": sample.provider,
                "freq": sample.freq,
                "rf_pwr": sample.rf_pwr,
                "comm_state": sample.comm_state,
                "user": sample.user,
                "station": sample.station,
                "captured_at_utc": sample.captured_at_utc,
                "device_id": sample.device_id,
                "received_at": sample.received_at.isoformat() if sample.received_at else None,
                "processed": sample.processed
            }
        return {"message": "No samples found"}
    except Exception as e:
        return {"error": str(e)}

# ==================== ERROR HANDLERS ====================

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    logger.error(f"HTTP error {exc.status_code}: {exc.detail}")
    return JSONResponse(
        status_code=exc.status_code,
        content={
            "status": "error",
            "message": exc.detail,
            "timestamp": time.time()
        }
    )

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled exception: {str(exc)}", exc_info=True)
    return JSONResponse(
        status_code=500,
        content={
            "status": "error", 
            "message": "Internal server error",
            "timestamp": time.time()
        }
    )

# ==================== REPORT ENDPOINTS ====================

@app.get("/report/csv")
async def generate_csv_report(
    device_id: str = "",
    start_date: str = "",
    end_date: str = "",
    db: Session = Depends(get_db)
):
    """
    Generate CSV report for location samples
    """
    try:
        # Build query
        query = db.query(models.LocationSample)
        
        if device_id:
            query = query.filter(models.LocationSample.device_id == device_id)
        
        if start_date:
            query = query.filter(models.LocationSample.sample_date >= start_date)
        
        if end_date:
            query = query.filter(models.LocationSample.sample_date <= end_date)
        
        # Get samples ordered by date and time
        samples = query.order_by(
            models.LocationSample.sample_date,
            models.LocationSample.sample_time
        ).all()
        
        # Generate CSV content WITHOUT accuracy and device ID
        csv_content = "Date,Time,Frequency,RF Power Output,Comm Result,Latitude,Longitude\n"
        
        for sample in samples:
            csv_content += f'"{sample.sample_date}","{sample.sample_time}","{sample.freq}","{sample.rf_pwr}","{sample.comm_state}",{sample.lat},{sample.lon}\n'
        
        return {
            "status": "success",
            "total_samples": len(samples),
            "csv_data": csv_content,
            "generated_at": time.time()
        }
        
    except Exception as e:
        logger.error(f"Report generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

@app.get("/report/summary")
async def get_report_summary(
    device_id: str = "",
    db: Session = Depends(get_db)
):
    """
    Get summary statistics for reporting
    """
    try:
        query = db.query(models.LocationSample)
        
        if device_id:
            query = query.filter(models.LocationSample.device_id == device_id)
        
        total_samples = query.count()
        
        # Get date range
        date_range = db.query(
            func.min(models.LocationSample.sample_date),
            func.max(models.LocationSample.sample_date)
        ).first()
        
        # Get unique frequencies
        frequencies = db.query(models.LocationSample.freq).distinct().all()
        frequencies = [f[0] for f in frequencies]
        
        return {
            "total_samples": total_samples,
            "date_range": {
                "start": date_range[0],
                "end": date_range[1]
            },
            "unique_frequencies": frequencies,
            "unique_devices": db.query(models.LocationSample.device_id).distinct().count()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Summary failed: {str(e)}")

# ==================== DASHBOARD ENDPOINTS ====================

@app.get("/api/trajectories")
async def get_trajectories(
    date_filter: Optional[str] = None,
    station_filter: Optional[str] = None,
    device_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get trajectory data with date filter and unique stations
    """
    try:
        # Use today's date if no date filter provided
        if date_filter:
            target_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
        else:
            target_date = datetime.now().date()

        # Build query using SQLAlchemy
        query = db.query(
            models.LocationSample.lat,
            models.LocationSample.lon,
            models.LocationSample.station,
            models.LocationSample.device_id,
            models.LocationSample.sample_date,
            models.LocationSample.sample_time,
            models.LocationSample.captured_at_utc,
            models.LocationSample.acc,
            models.LocationSample.received_at
        ).filter(
            models.LocationSample.sample_date == target_date.strftime("%Y-%m-%d")
        )
        
        # Apply filters
        if station_filter:
            query = query.filter(models.LocationSample.station == station_filter)
        
        if device_filter:
            query = query.filter(models.LocationSample.device_id == device_filter)
        
        # Execute query
        rows = query.distinct().order_by(
            models.LocationSample.station,
            models.LocationSample.device_id,
            models.LocationSample.captured_at_utc
        ).all()
        
        if not rows:
            return {
                "message": f"No data found for date {target_date}",
                "date_queried": target_date.isoformat(),
                "data": []
            }
        
        # Group coordinates by station and device_id
        trajectories = {}
        for row in rows:
            # Create a unique key for station + device combination
            trajectory_key = f"{row.station}_{row.device_id}"
            
            if trajectory_key not in trajectories:
                trajectories[trajectory_key] = {
                    'station': row.station,
                    'device_id': row.device_id,
                    'coordinates': []
                }
            
            # Create timestamp from sample_date and sample_time
            try:
                datetime_str = f"{row.sample_date} {row.sample_time}"
                timestamp = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S").isoformat()
            except:
                timestamp = f"{row.sample_date} {row.sample_time}"
            
            trajectories[trajectory_key]['coordinates'].append({
                'lat': float(row.lat),
                'lng': float(row.lon),
                'station': row.station,
                'device_id': row.device_id,
                'sample_date': row.sample_date,
                'sample_time': row.sample_time,
                'captured_at_utc': row.captured_at_utc,
                'accuracy': float(row.acc),
                'timestamp': timestamp
            })
        
        return {
            "date_queried": target_date.isoformat(),
            "total_trajectories": len(trajectories),
            "total_coordinates": sum(len(traj['coordinates']) for traj in trajectories.values()),
            "data": list(trajectories.values())
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"Error in get_trajectories: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/trajectories/date-range")
async def get_trajectories_date_range(
    start_date: str,
    end_date: str,
    station: Optional[str] = None,
    device_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get trajectory data for a date range
    """
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()

        # Build query
        query = db.query(
            models.LocationSample.lat,
            models.LocationSample.lon,
            models.LocationSample.station,
            models.LocationSample.device_id,
            models.LocationSample.sample_date,
            models.LocationSample.sample_time,
            models.LocationSample.captured_at_utc,
            models.LocationSample.acc,
            models.LocationSample.received_at
        ).filter(
            models.LocationSample.sample_date.between(
                start.strftime("%Y-%m-%d"), 
                end.strftime("%Y-%m-%d")
            )
        )
        
        if station:
            query = query.filter(models.LocationSample.station == station)
        
        if device_id:
            query = query.filter(models.LocationSample.device_id == device_id)
        
        rows = query.distinct().order_by(
            models.LocationSample.sample_date,
            models.LocationSample.station,
            models.LocationSample.device_id,
            models.LocationSample.captured_at_utc
        ).all()
        
        # Group by station and device
        trajectories = {}
        for row in rows:
            trajectory_key = f"{row.station}_{row.device_id}"
            
            if trajectory_key not in trajectories:
                trajectories[trajectory_key] = {
                    'station': row.station,
                    'device_id': row.device_id,
                    'coordinates': []
                }
            
            try:
                datetime_str = f"{row.sample_date} {row.sample_time}"
                timestamp = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S").isoformat()
            except:
                timestamp = f"{row.sample_date} {row.sample_time}"
            
            trajectories[trajectory_key]['coordinates'].append({
                'lat': float(row.lat),
                'lng': float(row.lon),
                'station': row.station,
                'device_id': row.device_id,
                'sample_date': row.sample_date,
                'sample_time': row.sample_time,
                'captured_at_utc': row.captured_at_utc,
                'accuracy': float(row.acc),
                'timestamp': timestamp
            })
        
        return {
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "total_trajectories": len(trajectories),
            "total_coordinates": sum(len(traj['coordinates']) for traj in trajectories.values()),
            "data": list(trajectories.values())
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"Error in get_trajectories_date_range: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/stations")
async def get_stations(db: Session = Depends(get_db)):
    """
    Get list of all unique stations
    """
    try:
        stations = db.query(models.LocationSample.station)\
                    .distinct()\
                    .order_by(models.LocationSample.station)\
                    .all()
        stations_list = [station[0] for station in stations]
        return {"stations": stations_list}
    except Exception as e:
        logger.error(f"Error in get_stations: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/devices")
async def get_devices(db: Session = Depends(get_db)):
    """
    Get list of all unique devices
    """
    try:
        devices = db.query(models.LocationSample.device_id)\
                   .distinct()\
                   .order_by(models.LocationSample.device_id)\
                   .all()
        devices_list = [device[0] for device in devices]
        return {"devices": devices_list}
    except Exception as e:
        logger.error(f"Error in get_devices: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

@app.get("/api/available-dates")
async def get_available_dates(db: Session = Depends(get_db)):
    """
    Get list of all available dates in the database
    """
    try:
        dates = db.query(models.LocationSample.sample_date)\
                 .distinct()\
                 .order_by(models.LocationSample.sample_date.desc())\
                 .all()
        dates_list = [date[0] for date in dates]
        return {"available_dates": dates_list}
    except Exception as e:
        logger.error(f"Error in get_available_dates: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=3003, log_level="info")