# app/reports.py
from fastapi import HTTPException, Depends
from sqlalchemy.orm import Session
import logging
from datetime import datetime
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
from . import models
from .database import get_db

logger = logging.getLogger(__name__)

def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two points using Haversine formula"""
    R = 6371000  # Earth radius in meters
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = (math.sin(delta_lat/2) * math.sin(delta_lat/2) + 
         math.cos(lat1_rad) * math.cos(lat2_rad) * 
         math.sin(delta_lon/2) * math.sin(delta_lon/2))
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c

def time_difference(time1, time2):
    """Calculate time difference in seconds between two HH:MM:SS strings"""
    try:
        # Handle cases where time might have milliseconds
        time1 = str(time1).split('.')[0]  # Remove milliseconds if present
        time2 = str(time2).split('.')[0]  # Remove milliseconds if present
        
        h1, m1, s1 = map(int, time1.split(':'))
        h2, m2, s2 = map(int, time2.split(':'))
        
        total_seconds1 = h1 * 3600 + m1 * 60 + s1
        total_seconds2 = h2 * 3600 + m2 * 60 + s2
        
        return abs(total_seconds1 - total_seconds2)
    except Exception as e:
        logger.warning(f"Error calculating time difference between {time1} and {time2}: {e}")
        return float('inf')  # Return a large value so they don't match

def get_comm_state_value(comm_state):
    """Convert comm_state string to numeric value"""
    comm_state = str(comm_state).strip().lower()
    if 'loud and clear' in comm_state:
        return 3
    elif 'readable noisy' in comm_state:
        return 2
    elif 'noisy' in comm_state:
        return 1
    else:
        return 0
    
def safe_filename(text):
    """Convert text to ASCII-safe filename"""
    # Replace non-ASCII characters with underscores or remove them
    text = re.sub(r'[^\x00-\x7F]+', '_', text)
    # Remove any other problematic characters
    text = re.sub(r'[<>:"/\\|?*]', '_', text)
    return text

def get_comm_state_style(comm_state_value):
    """Get Excel style based on comm_state value"""
    if comm_state_value == 3:
        return PatternFill(start_color="006400", end_color="006400", fill_type="solid"), Font(color="FFFFFF")  # Deep green, white text
    elif comm_state_value == 2:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"), Font(color="000000")  # Light green, black text
    elif comm_state_value == 1:
        return PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"), Font(color="000000")  # Orange, black text
    else:
        return PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"), Font(color="FFFFFF")  # Blood red, white text

async def generate_station_report(
    start_date: str,
    end_date: str,
    station1: str,
    station2: str,
    db: Session 
):
    """
    Generate professional Excel report for station pair analysis
    """
    try:
        logger.info(f"Generating report for {station1} and {station2} from {start_date} to {end_date}")
        
        # Validate dates
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError as e:
            logger.error(f"Date validation failed: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
        
        # Get data for both stations - use string dates for database comparison
        query = db.query(models.LocationSample).filter(
            models.LocationSample.sample_date.between(start_date, end_date),
            models.LocationSample.station.in_([station1, station2])
        ).order_by(
            models.LocationSample.sample_date,
            models.LocationSample.sample_time
        )
        
        samples = query.all()
        
        logger.info(f"Found {len(samples)} total samples")
        
        if not samples:
            raise HTTPException(status_code=404, detail="No data found for the specified criteria")
        
        # Separate samples by station - convert to strings for comparison
        station1_samples = [s for s in samples if str(s.station) == station1]
        station2_samples = [s for s in samples if str(s.station) == station2]
        
        logger.info(f"Station {station1}: {len(station1_samples)} samples")
        logger.info(f"Station {station2}: {len(station2_samples)} samples")
        
        if not station1_samples or not station2_samples:
            raise HTTPException(status_code=404, detail=f"No data found for stations: {station1} and/or {station2}")
        
        # Find matching time pairs (within 1.5 minutes = 90 seconds)
        matched_pairs = []
        
        logger.info("Starting time matching...")
        
        for s1 in station1_samples:
            for s2 in station2_samples:
                # Convert SQLAlchemy objects to strings for comparison
                s1_date = str(s1.sample_date)
                s2_date = str(s2.sample_date)
                s1_time = str(s1.sample_time)
                s2_time = str(s2.sample_time)
                
                time_diff = time_difference(s1_time, s2_time)
                
                if (s1_date == s2_date and time_diff <= 120):  # 1.5 minutes
                    logger.info(f"Found time match: {s1_time} vs {s2_time} (diff: {time_diff}s)")
                    
                    try:
                        # FIX: Access attributes directly without float() conversion
                        distance = haversine_distance(s1.lat, s1.lon, s2.lat, s2.lon)
                        comm_state_value = get_comm_state_value(s1.comm_state)
                        
                        matched_pairs.append({
                            'serial': len(matched_pairs) + 1,
                            'date': s1_date,
                            'time': f"{s1_time} (mean)",
                            'frequency': str(s1.freq),
                            'rf_power': str(s1.rf_pwr),
                            'lat_station1': s1.lat,
                            'lon_station1': s1.lon,
                            'lat_station2': s2.lat,
                            'lon_station2': s2.lon,
                            'distance': distance,
                            'comm_state': str(s1.comm_state),
                            'comm_state_value': comm_state_value
                        })
                        break  # Only match with first found pair
                    except Exception as e:
                        logger.error(f"Error processing matched pair: {e}")
                        continue
        
        logger.info(f"Found {len(matched_pairs)} matched pairs")
        
        if not matched_pairs:
            # Return a more helpful error message
            station1_times = [f"{s.sample_time}" for s in station1_samples]
            station2_times = [f"{s.sample_time}" for s in station2_samples]
            raise HTTPException(
                status_code=404, 
                detail=f"No matching time pairs found within 1.5-minute window. Station1 times: {station1_times}, Station2 times: {station2_times}"
            )
        
        # Create Excel workbook
        logger.info("Creating Excel workbook...")
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Radio Test Report")
        else:
            ws.title = "Radio Test Report"        
        # Define styles
        header_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title and header section - FIXED: Proper cell assignment
        ws.merge_cells('A1:K1')
        title_cell = ws.cell(row=1, column=1, value="Radio Set Field Test and Trial/ Field Functional Test Report - Auto Generated")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align
        
        # Header information
        ws.cell(row=3, column=1, value="Date:")
        date_cell = ws.cell(row=3, column=3, value=start_date)
        if start_date != end_date:
            date_cell.value = f"{start_date} to {end_date}"
        
        # Get time range
        times = [pair['time'].replace(' (mean)', '') for pair in matched_pairs]
        if times:
            ws.cell(row=4, column=1, value="Period Covering:")
            ws.cell(row=4, column=3, value=f"{min(times)} (initial time), {max(times)} (last time)")
        
        ws.cell(row=5, column=1, value="Stations:")
        ws.cell(row=5, column=3, value=f"{station1}, {station2}")
        
        ws.cell(row=6, column=1, value="Terrain Type:")
        ws.cell(row=6, column=3, value="----------Blank----------")
        
        # Table headers
        headers = ['Serial', 'Date', 'Time', 'Frequency', 'RF Power', 
                  f'Lat ({station1})', f'Lon ({station1})', 
                  f'Lat ({station2})', f'Lon ({station2})', 
                  'Distance (m)', 'Comm State']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Data rows - FIXED: Proper cell creation
        successful_communications = 0
        for row_idx, pair in enumerate(matched_pairs, 9):
            # Create all cells first
            ws.cell(row=row_idx, column=1, value=pair['serial'])
            ws.cell(row=row_idx, column=2, value=pair['date'])
            ws.cell(row=row_idx, column=3, value=pair['time'])
            ws.cell(row=row_idx, column=4, value=pair['frequency'])
            ws.cell(row=row_idx, column=5, value=pair['rf_power'])
            ws.cell(row=row_idx, column=6, value=round(pair['lat_station1'], 6))
            ws.cell(row=row_idx, column=7, value=round(pair['lon_station1'], 6))
            ws.cell(row=row_idx, column=8, value=round(pair['lat_station2'], 6))
            ws.cell(row=row_idx, column=9, value=round(pair['lon_station2'], 6))
            ws.cell(row=row_idx, column=10, value=round(pair['distance'], 2))
            
            # Communication state cell with styling
            comm_cell = ws.cell(row=row_idx, column=11, value=pair['comm_state'])
            fill, font = get_comm_state_style(pair['comm_state_value'])
            comm_cell.fill = fill
            comm_cell.font = font
            
            # Apply border to all cells in this row
            for col in range(1, 12):  # Columns 1 to 11
                ws.cell(row=row_idx, column=col).border = border
            
            # Count successful communications
            if pair['comm_state_value'] in [2, 3]:  # Readable Noisy or Loud and Clear
                successful_communications += 1
        
        # Summary section
        summary_row = len(matched_pairs) + 10
        ws.merge_cells(f'A{summary_row}:E{summary_row}')
        summary_cell = ws.cell(row=summary_row, column=1, value=f"Success Rate: {successful_communications}/{len(matched_pairs)} ({successful_communications/len(matched_pairs)*100:.1f}%)")
        summary_cell.font = Font(bold=True)
        summary_cell.alignment = center_align
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to bytes
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        # Generate filename
        # Use safe filenames
        safe_station1 = safe_filename(station1)
        safe_station2 = safe_filename(station2)
        filename = f"radio_report_{safe_station1}_{safe_station2}_{start_date}_{end_date}.xlsx"
        
        return {
            "status": "success",
            "message": f"Report generated with {len(matched_pairs)} matched pairs",
            "filename": filename,
            "file_data": file_stream.getvalue()
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Report generation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")